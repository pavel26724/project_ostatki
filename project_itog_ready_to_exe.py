import openpyxl
from openpyxl import load_workbook
import time
import datetime

def ostatok_base_44_ixora_44dealer_46alyans():
    book = load_workbook('остатки 1с.xlsx')
    book_price = load_workbook('price_ads.xlsx')
    book_const = load_workbook('КОНСТАНТА.xlsx')
    book_rez = openpyxl.Workbook()
    book_al = openpyxl.Workbook()
    book_44_dealer = openpyxl.Workbook()
    sheet = book.active
    sheet_price = book_price.active
    sheet_44_dealer = book_44_dealer.active
    sheet_rez = book_rez.active
    sheet_al = book_al.active
    sheet_const = book_const.active

    sheet_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    sheet_al.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    sheet_44_dealer.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    dict_ost = {}
    for row in sheet.iter_rows(min_row=11, max_row=None):
        if row[2].value != 0 and row[1].value:
            if row[2].value > 0:
                dict_ost[str(row[1].value)] = row[2].value

    dict_price = {}
    for row in sheet_price.iter_rows(min_row=2, max_row=None):
        a = dict(
            number=str(row[0].value),
            name=row[1].value,
            brand=row[2].value,
            partia=row[3].value,
            price_opt=row[4].value,
            price_uch=row[5].value,
            price_rrp=row[6].value,
        )
        dict_price[str(row[0].value)] = a

    for k, v in dict_ost.items():  # 44 и 44дилер
        k = str(k)
        if dict_price.get(k) and dict_price[k]['brand'] != 'BOSCH':
            a = [dict_price[k]['name'], dict_price[k]['brand'], dict_price[k]['number'], v + 2, dict_price[k]['partia'], dict_price[k]['price_rrp']]
            b = [dict_price[k]['name'], dict_price[k]['brand'], dict_price[k]['number'], v, dict_price[k]['partia'], dict_price[k]['price_opt']]
            sheet_44_dealer.append(a)
            sheet_rez.append(b)
        elif dict_price.get(k.rjust(10, '0')) and dict_price[k.rjust(10, '0')]['brand'] == 'BOSCH' and v <= 500:
            k = k.rjust(10, '0')
            a = [dict_price[k]['name'], dict_price[k]['brand'], dict_price[k]['number'], v + 2, dict_price[k]['partia'], dict_price[k]['price_opt']]
            b = [dict_price[k]['name'], dict_price[k]['brand'], dict_price[k]['number'], v, dict_price[k]['partia'], dict_price[k]['price_opt']]
            sheet_44_dealer.append(a)
            sheet_rez.append(b)
        elif dict_price.get(k.rjust(10, '0')) and dict_price[k.rjust(10, '0')]['brand'] == 'BOSCH':
            k = k.rjust(10, '0')
            b = [dict_price[k]['name'], dict_price[k]['brand'], dict_price[k]['number'], 1000, dict_price[k]['partia'], dict_price[k]['price_opt']]
            sheet_44_dealer.append(b)
            sheet_rez.append(b)

    for row in range(1, sheet_const.max_row):  # добаква константы
        name = sheet_const[row][0].value
        brand = sheet_const[row][1].value
        art = str(sheet_const[row][2].value)
        ost = int(sheet_const[row][3].value)
        partia = int(sheet_const[row][4].value)
        price_opt = sheet_const[row][5].value
        a = [name, brand, art, ost, partia, price_opt]
        sheet_rez.append(a)
        sheet_44_dealer.append(a)

    for k, v in dict_ost.items():  # 46 alyans
        k = str(k)
        if dict_price.get(k) and dict_price[k]['brand'] == 'SWF' or dict_price.get(k) and dict_price[k][
            'brand'] == 'MOTUL' or dict_price.get(k) and dict_price[k]['brand'] == 'Mando' or dict_price.get(k) and \
                dict_price[k]['brand'] == 'BSG' or dict_price.get(k) and dict_price[k]['brand'] == 'NIBK':
            b = [dict_price[k]['name'], dict_price[k]['brand'], dict_price[k]['number'], v, dict_price[k]['partia'], dict_price[k]['price_uch']]
            sheet_al.append(b)
        elif dict_price.get(k.rjust(10, '0')) and dict_price[k.rjust(10, '0')]['brand'] == 'BOSCH':
            k = k.rjust(10, '0')
            b = [dict_price[k]['name'], dict_price[k]['brand'], dict_price[k]['number'], v, dict_price[k]['partia'], dict_price[k]['price_uch']]
            sheet_al.append(b)

    book_rez.save(f'44/Остатки {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
    book_rez.save(f'44/IXORA.xlsx')
    book_rez.save(f'ost_base.xlsx')
    book_al.save(f'46учетАльянс/Остатки {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
    book_44_dealer.save(f'44дилер/Остатки {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
    book.close()
    book_rez.close()
    book_al.close()
    book_price.close()
    book_44_dealer.close()

def ostatok_46_46dealer_46BOSCH():
    book_ost = load_workbook('ost_base.xlsx')
    book_price = load_workbook('price_ads.xlsx')
    book_46rez = openpyxl.Workbook()
    book_46_bosch_rez = openpyxl.Workbook()
    book_46_dealer_rez = openpyxl.Workbook()
    book_44_bosch_rez = openpyxl.Workbook()
    sheet_46rez = book_46rez.active
    sheet_44_bosch_rez = book_44_bosch_rez.active
    sheet_ost = book_ost.active
    sheet_price = book_price.active
    sheet_46_dealer_rez = book_46_dealer_rez.active
    sheet_46_bosch_rez = book_46_bosch_rez.active

    sheet_46rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    sheet_46_bosch_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    sheet_46_dealer_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    sheet_44_bosch_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])

    dict_ost = []
    for row in sheet_ost.iter_rows(min_row=2, max_row=None):
        number=str(row[2].value)
        name=str(row[0].value)
        brand=str(row[1].value)
        partia=int(row[4].value)
        price_opt=float(row[5].value)
        ost=int(row[3].value)
        a = [name, brand, number, ost, partia, price_opt]
        dict_ost.append(a)


    for v in dict_ost:  # 46 и 46бош и 46дилер
        if v[1] != 'MOTUL':
            d = [v[0], v[1], v[2], v[3], v[4], v[5]]
            sheet_44_bosch_rez.append(d)
        if v[1] == 'MOTUL':
            a = [v[0][:28].upper(), v[1], v[2], '36', v[4], v[5]]
            b = [str(v[0][:28].upper()), str(v[1]), str(v[2]), str('22'), str(v[4]), str(v[5]).replace('.', ',')]
            c = [v[0][:28].upper(), v[1], v[2], '24', v[4], v[5]]
            sheet_46rez.append(a)
            sheet_46_bosch_rez.append(b)
            sheet_46_dealer_rez.append(c)

        else:
            if int(v[3]) <= 8:
                a = [v[0][:28].upper(), v[1], v[2], '8', v[4], v[5]]
                b = [str(v[0][:28].upper()), str(v[1]), str(v[2]), str('6'), str(v[4]), str(v[5]).replace('.', ',')]
                c = [v[0][:28].upper(), v[1], v[2], '10', v[4], v[5]]
                sheet_46rez.append(a)
                sheet_46_bosch_rez.append(b)
                sheet_46_dealer_rez.append(c)
            elif 8 < int(v[3]) <= 20:
                a = [v[0][:28].upper(), v[1], v[2], '20', v[4], v[5]]
                b = [str(v[0][:28].upper()), str(v[1]), str(v[2]), str('40'), str(v[4]), str(v[5]).replace('.', ',')]
                c = [v[0][:28].upper(), v[1], v[2], '30', v[4], v[5]]
                sheet_46rez.append(a)
                sheet_46_bosch_rez.append(b)
                sheet_46_dealer_rez.append(c)
            elif 20 < int(v[3]) <= 50:
                a = [v[0][:28].upper(), v[1], v[2], '50', v[4], v[5]]
                b = [str(v[0][:28].upper()), str(v[1]), str(v[2]), str('40'), str(v[4]), str(v[5]).replace('.', ',')]
                c = [v[0][:28].upper(), v[1], v[2], '40', v[4], v[5]]
                sheet_46rez.append(a)
                sheet_46_bosch_rez.append(b)
                sheet_46_dealer_rez.append(c)
            elif 50 < int(v[3]) <= 100:
                a = [v[0][:28].upper(), v[1], v[2], '100', v[4], v[5]]
                b = [str(v[0][:28].upper()), str(v[1]), str(v[2]), str('120'), str(v[4]), str(v[5]).replace('.', ',')]
                c = [v[0][:28].upper(), v[1], v[2], '90', v[4], v[5]]
                sheet_46rez.append(a)
                sheet_46_bosch_rez.append(b)
                sheet_46_dealer_rez.append(c)
            elif 100 < int(v[3]) <= 500:
                a = [v[0][:28].upper(), v[1], v[2], '500', v[4], v[5]]
                b = [str(v[0][:28].upper()), str(v[1]), str(v[2]), str('400'), str(v[4]), str(v[5]).replace('.', ',')]
                c = [v[0][:28].upper(), v[1], v[2], '300', v[4], v[5]]
                sheet_46rez.append(a)
                sheet_46_bosch_rez.append(b)
                sheet_46_dealer_rez.append(c)
            else:
                a = [v[0][:28].upper(), v[1], v[2], '1000', v[4], v[5]]
                b = [str(v[0][:28].upper()), str(v[1]), str(v[2]), str('800'), str(v[4]), str(v[5]).replace('.', ',')]
                c = [v[0][:28].upper(), v[1], v[2], '1200', v[4], v[5]]
                sheet_46rez.append(a)
                sheet_46_bosch_rez.append(b)
                sheet_46_dealer_rez.append(c)

    book_46rez.save(f'46/Остатки {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
    book_46_bosch_rez.save(f'46/BOSCH.xlsx')
    book_46_dealer_rez.save(f'46склад/Остатки {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
    book_44_bosch_rez.save(f'44/BOSCH.xlsx')
    book_46rez.close()
    book_46_bosch_rez.close()
    book_46_dealer_rez.close()
    book_44_bosch_rez.close()

start_time = time.time()
ostatok_base_44_ixora_44dealer_46alyans()
ostatok_46_46dealer_46BOSCH()
print(f'отработала за {int(time.time() - start_time)} секунд')

