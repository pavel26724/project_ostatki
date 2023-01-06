import os
import openpyxl
from openpyxl import load_workbook
import time
import datetime

_BSDIR = os.path.dirname(os.path.abspath(__file__))

def ostatok_base_44_44dealer_46alyans():
    book = load_workbook(os.path.join(_BSDIR, 'остатки 1с.xlsx'))
    book_price = load_workbook(os.path.join(_BSDIR, 'price_ads.xlsx'))
    book_const = load_workbook(os.path.join(_BSDIR, 'КОНСТАНТА.xlsx'))
    book_rez = openpyxl.Workbook()
    book_al = openpyxl.Workbook()
    book_44_dealer = openpyxl.Workbook()
    sheet = book.active
    sheet_price = book_price.active
    sheet_44_dealer = book_44_dealer.active
    sheet_rez = book_rez.active
    sheet_al = book_al.active
    sheet_const = book_const.active

    sheet_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    sheet_al.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    sheet_44_dealer.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    dict_ost = {}
    for row in sheet.iter_rows(min_row=11, max_row=None):
        if row[2].value != 0 and row[1].value:
            dict_ost[str(row[1].value)] = row[2].value
    dict_price = {}
    for row in sheet_price.iter_rows(min_row=2, max_row=None):
        a = dict(
            number=str(row[0].value),
            name=row[1].value,
            brand=row[2].value,
            partia=row[3].value,
            price_opt=row[4].value,
            price_uch=row[5].value,
            price_rrp=row[6].value,
        )
        dict_price[str(row[0].value)] = a

    for k, v in dict_ost.items():  # 44 и 44дилер
        k = str(k)
        if dict_price.get(k) and dict_price[k]['brand'] != 'BOSCH':
            a = [dict_price[k]['name'],
                 dict_price[k]['brand'],
                 dict_price[k]['number'], v + 2,
                 dict_price[k]['partia'],
                 dict_price[k]['price_rrp']]
            b = [dict_price[k]['name'],
                 dict_price[k]['brand'],
                 dict_price[k]['number'], v,
                 dict_price[k]['partia'],
                 dict_price[k]['price_opt']]
            sheet_44_dealer.append(a)
            sheet_rez.append(b)
        elif dict_price.get(k.rjust(10, '0')) and dict_price[k.rjust(10, '0')]['brand'] == 'BOSCH' and v <= 500:
            k = k.rjust(10, '0')
            a = [dict_price[k]['name'],
                 dict_price[k]['brand'],
                 dict_price[k]['number'], v + 2,
                 dict_price[k]['partia'],
                 dict_price[k]['price_opt']]
            b = [dict_price[k]['name'],
                 dict_price[k]['brand'],
                 dict_price[k]['number'], v,
                 dict_price[k]['partia'],
                 dict_price[k]['price_opt']]
            sheet_44_dealer.append(a)
            sheet_rez.append(b)
        elif dict_price.get(k.rjust(10, '0')) and dict_price[k.rjust(10, '0')]['brand'] == 'BOSCH':
            k = k.rjust(10, '0')
            b = [dict_price[k]['name'],
                 dict_price[k]['brand'],
                 dict_price[k]['number'], 1000,
                 dict_price[k]['partia'],
                 dict_price[k]['price_opt']]
            sheet_44_dealer.append(b)
            sheet_rez.append(b)

    for row in range(1, sheet_const.max_row):  # добаква константы
        name = sheet_const[row][0].value
        brand = sheet_const[row][1].value
        art = str(sheet_const[row][2].value)
        ost = int(sheet_const[row][3].value)
        partia = int(sheet_const[row][4].value)
        price_opt = sheet_const[row][5].value
        a = [name, brand, art, ost, partia, price_opt]
        sheet_rez.append(a)
        sheet_44_dealer.append(a)

    for k, v in dict_ost.items():  # 46 alyans
        k = str(k)
        if dict_price.get(k) and dict_price[k]['brand'] == 'SWF' or dict_price.get(k) and dict_price[k][
            'brand'] == 'MOTUL' or dict_price.get(k) and dict_price[k]['brand'] == 'MANDO' or dict_price.get(k) and \
                dict_price[k]['brand'] == 'BSG':
            b = [dict_price[k]['name'],
                 dict_price[k]['brand'],
                 dict_price[k]['number'], v,
                 dict_price[k]['partia'],
                 dict_price[k]['price_uch']]
            sheet_al.append(b)
        elif dict_price.get(k.rjust(10, '0')) and dict_price[k.rjust(10, '0')]['brand'] == 'BOSCH':
            k = k.rjust(10, '0')
            b = [dict_price[k]['name'],
                 dict_price[k]['brand'],
                 dict_price[k]['number'], v,
                 dict_price[k]['partia'],
                 dict_price[k]['price_uch']]
            sheet_al.append(b)

    book_rez.save(f'44/Остатки {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
    book_rez.save(f'ost_base.xlsx')
    book_al.save(f'46учетАльянс/Остатки {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
    book_44_dealer.save(f'44дилер/Остатки {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
    book.close()
    book_rez.close()
    book_al.close()
    book_price.close()
    book_44_dealer.close()

def ostatok_46_46dealer_46BOSCH():
    book_ost = load_workbook(os.path.join(_BSDIR, 'ost_base.xlsx'))
    book_price = load_workbook(os.path.join(_BSDIR, 'price_ads.xlsx'))
    book_46rez = openpyxl.Workbook()
    book_46_bosch_rez = openpyxl.Workbook()
    book_46_dealer_rez = openpyxl.Workbook()
    book_44_bosch_rez = openpyxl.Workbook()
    sheet_46rez = book_46rez.active
    sheet_44_bosch_rez = book_44_bosch_rez.active
    sheet_ost = book_ost.active
    sheet_price = book_price.active
    sheet_46_dealer_rez = book_46_dealer_rez.active
    sheet_46_bosch_rez = book_46_bosch_rez.active

    sheet_46rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    sheet_46_bosch_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    sheet_46_dealer_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
    sheet_44_bosch_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])

    dict_ost = {}
    for row in sheet_ost.iter_rows(min_row=2, max_row=None):
        a = dict(
            number=row[2].value,
            name=row[0].value,
            brand=row[1].value,
            partia=row[4].value,
            price_opt=row[5].value,
            ost=row[3].value
        )
        dict_ost[row[2].value] = a
    dict_price = {}
    for row in sheet_price.iter_rows(min_row=2, max_row=None):
        a = dict(
            number=row[0].value,
            name=row[1].value,
            brand=row[2].value,
            partia=row[3].value,
            price_opt=row[4].value,
            price_uch=row[5].value,
            price_rrp=row[6].value
        )
        dict_price[row[0].value] = a

    for v in dict_ost.values():  # 46 и 46бош и 46дилер
        if v['ost'] != 'Остаток':
            if v['brand'] != 'MOTUL':
                d = [v['name'], v['brand'], v['number'], v['ost'], v['partia'], v['price_opt']]
                sheet_44_bosch_rez.append(d)
            if v['brand'] == 'MOTUL':
                a = [v['name'], v['brand'], v['number'], '36', v['partia'], v['price_opt']]
                b = [v['name'], v['brand'], v['number'], '22', v['partia'], v['price_opt']]
                c = [v['name'], v['brand'], v['number'], '24', v['partia'], v['price_opt']]
                sheet_46rez.append(a)
                sheet_46_bosch_rez.append(b)
                sheet_46_dealer_rez.append(c)

            else:
                if int(v['ost']) <= 8:
                    a = [v['name'], v['brand'], v['number'], '8', v['partia'], v['price_opt']]
                    b = [v['name'], v['brand'], v['number'], '6', v['partia'], v['price_opt']]
                    c = [v['name'], v['brand'], v['number'], '10', v['partia'], v['price_opt']]
                    sheet_46rez.append(a)
                    sheet_46_bosch_rez.append(b)
                    sheet_46_dealer_rez.append(c)
                elif int(v['ost']) > 8 and int(v['ost']) <= 20:
                    a = [v['name'], v['brand'], v['number'], '20', v['partia'], v['price_opt']]
                    b = [v['name'], v['brand'], v['number'], '40', v['partia'], v['price_opt']]
                    c = [v['name'], v['brand'], v['number'], '30', v['partia'], v['price_opt']]
                    sheet_46rez.append(a)
                    sheet_46_bosch_rez.append(b)
                    sheet_46_dealer_rez.append(c)
                elif int(v['ost']) > 20 and int(v['ost']) <= 50:
                    a = [v['name'], v['brand'], v['number'], '50', v['partia'], v['price_opt']]
                    b = [v['name'], v['brand'], v['number'], '40', v['partia'], v['price_opt']]
                    c = [v['name'], v['brand'], v['number'], '40', v['partia'], v['price_opt']]
                    sheet_46rez.append(a)
                    sheet_46_bosch_rez.append(b)
                    sheet_46_dealer_rez.append(c)
                elif int(v['ost']) > 50 and int(v['ost']) <= 100:
                    a = [v['name'], v['brand'], v['number'], '100', v['partia'], v['price_opt']]
                    b = [v['name'], v['brand'], v['number'], '120', v['partia'], v['price_opt']]
                    c = [v['name'], v['brand'], v['number'], '90', v['partia'], v['price_opt']]
                    sheet_46rez.append(a)
                    sheet_46_bosch_rez.append(b)
                    sheet_46_dealer_rez.append(c)
                elif int(v['ost']) > 100 and int(v['ost']) <= 500:
                    a = [v['name'], v['brand'], v['number'], '500', v['partia'], v['price_opt']]
                    b = [v['name'], v['brand'], v['number'], '400', v['partia'], v['price_opt']]
                    c = [v['name'], v['brand'], v['number'], '300', v['partia'], v['price_opt']]
                    sheet_46rez.append(a)
                    sheet_46_bosch_rez.append(b)
                    sheet_46_dealer_rez.append(c)
                else:
                    a = [v['name'], v['brand'], v['number'], '1000', v['partia'], v['price_opt']]
                    b = [v['name'], v['brand'], v['number'], '1000', v['partia'], v['price_opt']]
                    c = [v['name'], v['brand'], v['number'], '1200', v['partia'], v['price_opt']]
                    sheet_46rez.append(a)
                    sheet_46_bosch_rez.append(b)
                    sheet_46_dealer_rez.append(c)

    book_46rez.save(f'46/Остатки {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
    book_46_bosch_rez.save(f'46/BOSCH.xlsx')
    book_46_dealer_rez.save(f'46дилер/Остатки {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
    book_44_bosch_rez.save(f'44/BOSCH.xlsx')
    book_46rez.close()
    book_46_bosch_rez.close()
    book_46_dealer_rez.close()
    book_44_bosch_rez.close()

start_time = time.time()
ostatok_base_44_44dealer_46alyans()
ostatok_46_46dealer_46BOSCH()
print(f'отработла за {int(time.time() - start_time)} секунд')
